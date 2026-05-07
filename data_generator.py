import random
import pandas as pd
import numpy as np
from faker import Faker
from datetime import date, timedelta
import uuid
import warnings
import os

warnings.filterwarnings('ignore')

fake = Faker()
random.seed(42)
np.random.seed(42)
Faker.seed(42)

# ============================================================================
# 1. Configuration & Reference Tables
# ============================================================================
N = 50_000  # number of claims / policies

STATES = ['NY','CA','TX','FL','IL','PA','OH','GA','NC','MI','NJ','VA','WA','AZ','MA']
MAKES_MODELS = {
    'Toyota':     ['Camry','Corolla','RAV4','Highlander','Tacoma'],
    'Honda':      ['Civic','Accord','CR-V','Pilot','Odyssey'],
    'Ford':       ['F-150','Fusion','Explorer','Escape','Mustang'],
    'Chevrolet':  ['Silverado','Malibu','Equinox','Tahoe','Traverse'],
    'BMW':        ['3 Series','5 Series','X5','X3','7 Series'],
    'Nissan':     ['Altima','Sentra','Rogue','Pathfinder','Murano'],
    'Jeep':       ['Grand Cherokee','Wrangler','Cherokee','Compass','Gladiator'],
    'Dodge':      ['Charger','Challenger','Durango','Ram 1500','Journey'],
    'Hyundai':    ['Elantra','Sonata','Tucson','Santa Fe','Kona'],
    'Kia':        ['Optima','Sorento','Sportage','Soul','Telluride'],
}
OCCUPATIONS   = ['exec-managerial','tech-support','craft-repair','other-service',
                  'sales','armed-forces','prof-specialty','transport-moving',
                  'farming-fishing','machine-op-inspct','adm-clerical','handlers-cleaners']
HOBBIES       = ['camping','chess','exercise','golf','video-games','reading','hiking',
                  'hunting','skydiving','paintball','polo','kayaking','sleeping','dancing','yoga']
EDUCATION     = ['High School','Bachelors','Masters','PhD','MD','JD','Associate']
RELATIONSHIP  = ['husband','wife','own-child','unmarried','other-relative','not-in-family']
INC_TYPES     = ['Single Vehicle Collision','Multi-vehicle Collision',
                  'Parked Car','Vehicle Theft','Other']
COL_TYPES     = ['Front Collision','Rear Collision','Side Collision','?']
SEVERITIES    = ['Minor Damage','Major Damage','Total Loss','Trivial Damage']
AUTHORITIES   = ['Police','Fire','Ambulance','None','Other']
LOCATIONS     = ['Highway','Parking Lot','Local Road','Interstate','Residential Street']
CSLS          = ['100/300','250/500','500/1000']
DEDUCTIBLES   = [250, 500, 1000, 2000]
COVERAGES     = {
    'LIAB':     'Liability',
    'COMP':     'Comprehensive',
    'COLL':     'Collision',
    'UNINSMOT': 'Uninsured Motorist',
    'MED':      'Medical Payments',
    'PIP':      'Personal Injury Protection',
}
POLICY_TYPES  = ['Auto Comprehensive','Auto Collision','Auto Liability','Full Coverage']
PAY_STATUS    = ['Current','Delinquent','Cancelled','Lapsed','Paid in Full']
QUOTE_STATUS  = ['Bound','Declined','Pending','Expired','Quoted']
RISK_CLASSES  = ['Preferred','Standard','Non-Standard','High-Risk']
VEH_USES      = ['Commute','Pleasure','Business','Farm','Rideshare']
INCOME_RNGS   = ['<$25K','$25K-$50K','$50K-$75K','$75K-$100K','$100K-$150K','>$150K']
INSURERS      = ['State Farm','GEICO','Allstate','Progressive','Farmers','USAA','None']
DECLINE_REASONS = [
    'Coverage lapsed at incident date',
    'Incident type excluded by policy',
    'Claim amount exceeds coverage limit',
    'Fraud score above threshold',
    'Policy cancelled prior to loss',
    'Vehicle not listed on policy',
    'Driver not listed on policy',
    'Duplicate claim detected',
    'Policy not active at time of loss',
    'Coverage limit exhausted by prior claims',
]
FRAUD_PATTERNS = [
    'Staged collision with known associate',
    'Inflated repair estimate submitted',
    'Phantom passenger injury claimed',
    'Vehicle reported stolen but located nearby',
    'Incident date outside policy period',
    'Multiple claims same vehicle short period',
    'Repair shop flagged for collusion',
    'Inconsistent witness statements',
]
SUBMISSION_TEMPLATES = [
    "Hi, I'd like to report a {inc_type} that occurred on {inc_date} at {location} in {city}, {state}. My vehicle, a {year} {make} {model}, sustained {severity}. Policy number is {policy}.",
    "I need to file a claim. On {inc_date}, I was involved in a {inc_type} on {location} in {city}. The damage to my {year} {make} {model} is {severity}. Please process under policy {policy}.",
    "Claim submission: Date of loss {inc_date}. Type: {inc_type}. Location: {location}, {city}, {state}. Vehicle: {year} {make} {model}. Damage assessed as {severity}. Policy ref: {policy}.",
    "My car was in an accident on {inc_date}. It happened at {location} in {city}, {state}. I drive a {year} {make} {model} and the damage looks like {severity}. My policy is {policy}.",
    "Filing auto claim – {inc_date}, {inc_type}, {city} {state}. Vehicle {year} {make} {model} shows {severity}. Authorities contacted: {authorities}. Policy: {policy}. Please advise next steps.",
]
ADJUSTER_NOTE_TEMPLATES = [
    "Spoke with {name}. Claimant reports {description}. Recommended {action}. Follow-up scheduled for {date}.",
    "Field inspection completed at {location}. Observed {observations}. Photos taken. Estimate pending from {shop}.",
    "Medical records received from {provider}. Treating physician notes {findings}. {days} days of treatment documented. IME {ime_status}.",
    "Witness interview conducted with {witness_name}. Statement {consistency} claimant account. Additional verification {needed}.",
    "Vehicle inspection at {shop}. Initial estimate ${amount}. {parts_status} parts availability. Repair timeline {days} days. Alternative transportation {rental_status}.",
    "Reviewed prior claims history. {count} claims in {years} years. Pattern {pattern_status}. Referred to SIU for further review.",
    "Policy coverage verified. Incident date {coverage_status} within policy period. Coverage {limit_status} sufficient for claimed damages. Deductible ${deductible} applies.",
    "Contacted {authority} for incident report verification. Report #{report_num} {status}. Details {match_status} claimant statement.",
    "Claimant submitted {doc_type}. Document appears {authenticity}. {discrepancy_note}. Requesting additional {additional_docs}.",
    "Liability assessment: Based on {evidence}, liability appears {liability_pct}% on insured. Comparative negligence {negligence_note}.",
]
REPAIR_ESTIMATE_TEMPLATES = [
    """REPAIR ESTIMATE
Shop: {shop_name}
Date: {est_date}
Vehicle: {year} {make} {model}
VIN: {vin}
Odometer: {odometer} miles

Damage Assessment:
- {damage_1}: ${cost_1}
- {damage_2}: ${cost_2}
- {damage_3}: ${cost_3}
- Labor ({hours} hrs @ ${rate}/hr): ${labor_cost}

Parts Total: ${parts_total}
Labor Total: ${labor_total}
Tax: ${tax}
GRAND TOTAL: ${grand_total}

Estimated Completion: {completion_date}
Notes: {notes}""",
    """BODY SHOP ESTIMATE
Facility: {shop_name}
Estimator: {estimator_name}
Date Prepared: {est_date}

Vehicle Information:
{year} {make} {model} | VIN: {vin} | Mileage: {odometer}

Required Repairs:
1. {damage_1} — ${cost_1}
2. {damage_2} — ${cost_2}
3. {damage_3} — ${cost_3}

Subtotal Parts: ${parts_total}
Labor: {hours} hours x ${rate}/hour = ${labor_total}
Shop Supplies: ${supplies}
Tax: ${tax}

TOTAL ESTIMATE: ${grand_total}

Insurance Deductible: ${deductible}
Customer Responsibility: ${customer_pay}

Estimate Valid Until: {valid_until}""",
]
MEDICAL_REPORT_TEMPLATES = [
    """MEDICAL EVALUATION REPORT
Patient: {patient_name}
DOB: {dob}
Date of Exam: {exam_date}
Referring Physician: {referring_md}

History of Present Illness:
Patient presents following motor vehicle accident on {accident_date}. Reports {symptoms}.
Onset: {onset}. Pain level: {pain_level}/10.

Physical Examination:
- Cervical spine: {cervical_findings}
- Lumbar spine: {lumbar_findings}
- Range of motion: {rom_findings}
- Neurological: {neuro_findings}

Diagnostic Imaging:
{imaging_type} performed on {imaging_date}. Findings: {imaging_findings}

Diagnosis:
1. {diagnosis_1}
2. {diagnosis_2}

Treatment Plan:
- {treatment_1}
- {treatment_2}
- Follow-up: {follow_up_weeks} weeks
- Estimated recovery: {recovery_weeks} weeks

Disability Status: {disability_status}
Return to Work: {rtw_date}

Provider Signature: {provider_name}, {provider_credentials}""",
    """EMERGENCY DEPARTMENT RECORD
Facility: {facility_name}
Arrival: {arrival_time}
Triage Level: {triage_level}

Chief Complaint: {chief_complaint}

Mechanism of Injury: Motor vehicle collision, {collision_detail}

Vital Signs: BP {bp}, HR {hr}, RR {rr}, O2 Sat {o2}%

Assessment:
- Primary Survey: {primary_survey}
- Secondary Survey: {secondary_survey}

Procedures Performed:
{procedures}

Results:
- X-ray: {xray_result}
- CT: {ct_result}
- Labs: {lab_result}

Disposition: {disposition}
Discharge Instructions: {discharge_instructions}
Follow-up: {follow_up}

Attending: {attending_name}, MD""",
]

def rand_date(s, e):
    return s + timedelta(days=random.randint(0, (e - s).days))

def make_policy_number():
    return f"POL-{random.randint(100000,999999)}"

def make_customer_id():
    return f"CUST-{random.randint(10000,99999)}"

def make_claim_id():
    return f"CL-{random.randint(100000,999999)}"

def make_vehicle_id():
    return f"VEH-{random.randint(100000,999999)}"

# ============================================================================
# 2. Pre-generate shared keys
# ============================================================================
N_CUSTOMERS = int(N * 0.7)  # ~35,000 unique customers

customer_pool = [make_customer_id() for _ in range(N_CUSTOMERS)]
policy_pool   = [make_policy_number() for _ in range(N)]
claim_pool    = [make_claim_id() for _ in range(N)]
vehicle_pool  = [make_vehicle_id() for _ in range(N)]   # one vehicle per policy slot

# Map policy -> customer (some customers have multiple policies)
policy_customer_map = {}
for i in range(N):
    if i < N_CUSTOMERS:
        policy_customer_map[policy_pool[i]] = customer_pool[i]
    else:
        policy_customer_map[policy_pool[i]] = random.choice(customer_pool)

# Map policy -> vehicle (1:1 for simplicity; some customers share a vehicle)
policy_vehicle_map = {policy_pool[i]: vehicle_pool[i] for i in range(N)}

# Pre-generate vehicle specs (used by both Vehicle table and claims/docs)
make_pool  = random.choices(list(MAKES_MODELS.keys()), k=N)
model_pool = [random.choice(MAKES_MODELS[m]) for m in make_pool]
year_pool  = [random.randint(2004, 2023) for _ in range(N)]

fraud_pool = random.choices([True, False], weights=[0.24, 0.76], k=N)

# ============================================================================
# 3. Customer_Master
#    Owns: identity, demographics, financials, credit
#    Does NOT own: policy terms, vehicle specs, claim facts
# ============================================================================
print("Building Customer_Master ...")

customer_attributes = {}
for cust_id in customer_pool:
    customer_attributes[cust_id] = {
        'gender':              random.choice(['MALE', 'FEMALE']),
        'education_level':     random.choice(EDUCATION),
        'occupation':          random.choice(OCCUPATIONS),
        'hobbies':             random.choice(HOBBIES),
        'relationship_status': random.choice(RELATIONSHIP),
        'credit_score':        random.randint(450, 850),
        # Store dates, not derived numbers — age/tenure computed at query time
        'date_of_birth':       rand_date(date(1948, 1, 1), date(2005, 1, 1)),
        'customer_since_date': rand_date(date(2005, 1, 1), date(2022, 1, 1)),
        # Customer financial attributes — belong here, nowhere else
        'capital_gains':       random.choice([0]*7 + [random.randint(1000, 60000)]),
        'capital_loss':        random.choice([0]*8 + [random.randint(500, 25000)]),
    }

customer_master_rows = []
for cust_id in customer_pool:
    a = customer_attributes[cust_id]
    policies_for_cust = [p for p, c in policy_customer_map.items() if c == cust_id]
    customer_master_rows.append({
        'customer_id':          cust_id,
        'customer_name':        fake.name(),
        'date_of_birth':        a['date_of_birth'].strftime('%Y-%m-%d'),
        'customer_since_date':  a['customer_since_date'].strftime('%Y-%m-%d'),
        'customer_address':     fake.address().replace('\n', ', '),
        'customer_phone':       fake.phone_number(),
        'customer_email':       fake.email(),
        'gender':               a['gender'],
        'education_level':      a['education_level'],
        'occupation':           a['occupation'],
        'hobbies':              a['hobbies'],
        'relationship_status':  a['relationship_status'],
        'credit_score':         a['credit_score'],
        'capital_gains':        a['capital_gains'],
        'capital_loss':         a['capital_loss'],
        'customer_segment':     random.choice(['Standard','Preferred','High-Value','New Customer']),
        'total_policies':       len(policies_for_cust),
        'active_policies':      random.randint(0, max(1, len(policies_for_cust))),
        'lifetime_claims':      random.choices([0,1,2,3,4,5], weights=[0.4,0.25,0.16,0.10,0.06,0.03])[0],
    })

df_customer_master = pd.DataFrame(customer_master_rows)
print(f"  Customer_Master: {len(df_customer_master):,} rows, {len(df_customer_master.columns)} cols")

# ============================================================================
# 4. Vehicle
#    Owns: VIN, make, model, year, value, usage, telematics
#    Referenced by: Policy_Data (FK) and Claim (FK)
# ============================================================================
print("Building Vehicle ...")

vehicle_rows = []
for i in range(N):
    cust = policy_customer_map[policy_pool[i]]
    vehicle_rows.append({
        'vehicle_id':          vehicle_pool[i],
        'customer_id':         cust,                          # FK → Customer_Master
        'vin':                 fake.vin(),
        'make':                make_pool[i],
        'model':               model_pool[i],
        'year':                year_pool[i],
        'body_style':          random.choice(['Sedan','SUV','Truck','Coupe','Minivan','Hatchback']),
        'color':               random.choice(['White','Black','Silver','Gray','Red','Blue','Green']),
        'market_value':        round(random.uniform(4000, 72000), 2),
        'annual_mileage':      random.randint(3000, 35000),
        'primary_use':         random.choice(VEH_USES),
        'telematics_enrolled': random.choice(['Yes','No']),
        'telematics_score':    round(random.uniform(20, 100), 1),
    })

df_vehicle = pd.DataFrame(vehicle_rows).drop_duplicates(subset='vehicle_id')
print(f"  Vehicle: {len(df_vehicle):,} rows, {len(df_vehicle.columns)} cols")

# ============================================================================
# 5. Policy_Data
#    Owns: policy terms, dates, premium, payment status, umbrella limit
#    Does NOT own: customer demographics, vehicle specs, credit score
# ============================================================================
print("Building Policy_Data ...")

policy_rows = []
for i in range(N):
    eff        = rand_date(date(2016, 1, 1), date(2023, 6, 1))
    exp        = eff + timedelta(days=365)
    pol_status = random.choices(['Active','Expired','Cancelled'], [0.62, 0.28, 0.10])[0]
    cust       = policy_customer_map[policy_pool[i]]
    veh        = policy_vehicle_map[policy_pool[i]]

    policy_rows.append({
        'policy_number':      policy_pool[i],
        'customer_id':        cust,                           # FK → Customer_Master
        'vehicle_id':         veh,                            # FK → Vehicle
        'policy_type':        random.choice(POLICY_TYPES),
        'policy_status':      pol_status,
        'effective_date':     eff.strftime('%Y-%m-%d'),
        'expiration_date':    exp.strftime('%Y-%m-%d'),
        'bind_date':          (eff - timedelta(days=random.randint(1,30))).strftime('%Y-%m-%d'),
        'total_annual_premium': round(random.uniform(600, 4500), 2),
        'payment_status':     random.choice(PAY_STATUS),
        'umbrella_limit':     random.choice([0, 100000, 500000, 1000000, 2000000]),
        'insured_name':       fake.name(),
        'insured_address':    fake.address().replace('\n', ', '),
        'insured_phone':      fake.phone_number(),
        'insured_email':      fake.email(),
        'insured_state':      random.choice(STATES),
        'insured_zip':        fake.zipcode(),
        # credit_score intentionally NOT here — join from Customer_Master
    })

df_policy = pd.DataFrame(policy_rows)
print(f"  Policy_Data: {len(df_policy):,} rows, {len(df_policy.columns)} cols")

# ============================================================================
# 6. Coverage_Line
#    One row per coverage type per policy — owns limit, deductible, CSL, premium
#    This is where policy_csl, coverage_limit, coverage_deductible belong
# ============================================================================
print("Building Coverage_Line ...")

coverage_rows = []
cov_codes = list(COVERAGES.keys())
for i in range(N):
    # Every policy gets 2-4 coverage lines
    n_coverages = random.choices([2, 3, 4], weights=[0.3, 0.5, 0.2])[0]
    selected_codes = random.sample(cov_codes, n_coverages)
    base_premium = df_policy.iloc[i]['total_annual_premium']
    splits = [random.random() for _ in selected_codes]
    split_total = sum(splits)
    for j, cc in enumerate(selected_codes):
        coverage_rows.append({
            'coverage_id':       f"COV-{str(uuid.uuid4())[:8].upper()}",
            'policy_number':     policy_pool[i],              # FK → Policy_Data
            'coverage_code':     cc,
            'coverage_name':     COVERAGES[cc],
            'coverage_limit':    random.choice([50000, 100000, 250000, 500000, 1000000]),
            'deductible':        random.choice(DEDUCTIBLES),  # correct spelling
            'csl':               random.choice(CSLS),
            'premium_for_line':  round(base_premium * splits[j] / split_total, 2),
        })

df_coverage = pd.DataFrame(coverage_rows)
print(f"  Coverage_Line: {len(df_coverage):,} rows, {len(df_coverage.columns)} cols")

# ============================================================================
# 7. Claim
#    Owns: incident facts, amounts, FKs
#    Does NOT own: customer demographics, policy terms, vehicle specs,
#                  fraud scores, computed booleans, free-text notes
# ============================================================================
print("Building Claim ...")

claims = []
fraud_assessments = []
documents = []   # built alongside claims for submission text

for i in range(N):
    is_fraud   = fraud_pool[i]
    policy     = policy_pool[i]
    customer   = policy_customer_map[policy]
    vehicle_id = policy_vehicle_map[policy]
    claim_id   = claim_pool[i]
    make       = make_pool[i]
    model      = model_pool[i]
    yr         = year_pool[i]
    state      = random.choice(STATES)
    city       = fake.city()
    severity   = random.choice(SEVERITIES)
    inc_type   = random.choice(INC_TYPES)
    col_type   = random.choice(COL_TYPES)
    auth       = random.choice(AUTHORITIES)
    location   = random.choice(LOCATIONS)

    # Policy dates (looked up from policy table by index for generation speed)
    policy_eff = date.fromisoformat(df_policy.iloc[i]['effective_date'])
    policy_exp = date.fromisoformat(df_policy.iloc[i]['expiration_date'])
    inc        = rand_date(date(2019, 1, 1), date(2024, 6, 1))

    # Derived values — computed here for fraud_assessment table only
    incident_in_policy     = policy_eff <= inc <= policy_exp
    incident_near_boundary = (abs((inc - policy_eff).days) <= 30 or
                               abs((policy_exp - inc).days) <= 30)

    if inc > policy_exp:
        policy_status_at_incident = 'Expired'
    elif inc < policy_eff:
        policy_status_at_incident = 'Not Yet Effective'
    else:
        policy_status_at_incident = random.choices(
            ['Active','Active','Active','Cancelled','Lapsed'],
            weights=[0.70, 0.15, 0.05, 0.07, 0.03]
        )[0]

    # ---- Claim amounts & fraud logic ----
    if is_fraud:
        total   = random.randint(8000, 90000)
        hour    = random.choices(list(range(24)),
                    weights=[4,4,4,4,3,2,1,1,1,1,1,1,1,1,1,1,2,3,4,5,5,5,4,4])[0]
        witn    = random.choices([0,1,2,3], weights=[0.5,0.3,0.15,0.05])[0]
        prior_c = random.choices([0,1,2,3,4], weights=[0.1,0.2,0.3,0.25,0.15])[0]
        fraud_pattern  = random.choice(FRAUD_PATTERNS)
        fraud_score    = round(random.uniform(0.65, 0.99), 4)
        rule_flag      = random.choice(['HIGH_CLAIM','SUSPICIOUS_HOUR','PRIOR_CLAIMS',
                                         'WITNESS_MISMATCH','DUPLICATE_RISK'])
        if not incident_in_policy or policy_status_at_incident != 'Active':
            decline_reason = random.choice([
                'Policy not active at time of loss',
                'Coverage lapsed at incident date',
                'Policy cancelled prior to loss',
                'Incident date outside policy period',
            ])
        else:
            decline_reason = random.choice(DECLINE_REASONS) if random.random() < 0.5 else ''
        claim_outcome = random.choices(
            ['Denied','Settled - Reduced','Under Investigation'],
            weights=[0.45, 0.30, 0.25]
        )[0]
    else:
        total   = random.randint(300, 25000)
        hour    = random.choices(list(range(24)),
                    weights=[1,1,1,1,1,2,4,6,7,7,6,5,5,6,6,5,5,5,4,3,3,3,2,1])[0]
        witn    = random.choices([0,1,2,3], weights=[0.25,0.45,0.2,0.1])[0]
        prior_c = random.choices([0,1,2,3,4], weights=[0.5,0.28,0.13,0.06,0.03])[0]
        fraud_pattern  = ''
        fraud_score    = round(random.uniform(0.01, 0.35), 4)
        rule_flag      = random.choices(['NONE','LOW_SEVERITY','REVIEW_REQUIRED'],
                                         weights=[0.7, 0.2, 0.1])[0]
        if not incident_in_policy or policy_status_at_incident != 'Active':
            decline_reason = random.choice([
                'Policy not active at time of loss',
                'Coverage lapsed at incident date',
                'Policy cancelled prior to loss',
            ])
            claim_outcome = random.choices(['Denied','Closed - No Payment'], weights=[0.7, 0.3])[0]
        else:
            decline_reason = random.choice(DECLINE_REASONS) if random.random() < 0.05 else ''
            claim_outcome  = random.choices(
                ['Approved','Settled','Closed - No Payment'],
                weights=[0.6, 0.3, 0.1]
            )[0]

    injury = random.randint(0, int(total * 0.35)) if severity != 'Trivial Damage' else 0
    prop   = random.randint(0, int(total * 0.25))
    veh_cl = total - injury - prop
    bodily_injuries_val = random.randint(0, 3)

    is_complex = ((total > 15000) or is_fraud or (bodily_injuries_val > 1) or
                  not incident_in_policy or policy_status_at_incident != 'Active' or
                  random.random() < 0.15)

    # ---- CLAIM row — incident facts + FKs only ----
    claims.append({
        'claim_id':                   claim_id,
        'customer_id':                customer,          # FK → Customer_Master
        'policy_number':              policy,            # FK → Policy_Data
        'vehicle_id':                 vehicle_id,        # FK → Vehicle
        'incident_date':              inc.strftime('%Y-%m-%d'),
        'incident_type':              inc_type,
        'collision_type':             col_type,
        'incident_severity':          severity,
        'authorities_contacted':      auth,
        'incident_state':             state,
        'incident_city':              city,
        'incident_location':          location,
        'incident_hour_of_the_day':   hour,
        'number_of_vehicles_involved': random.randint(1, 4),
        'property_damage':            random.choice(['YES','NO','?']),
        'bodily_injuries':            bodily_injuries_val,
        'witnesses':                  witn,
        'police_report_available':    random.choice(['YES','NO','?']),
        'total_claim_amount':         total,
        'injury_claim':               injury,
        'property_claim':             prop,
        'vehicle_claim':              veh_cl,
        'prior_claims_count':         prior_c,
        'claim_status':               'Open' if claim_outcome in ['Under Investigation'] else 'Closed',
    })

    # ---- FRAUD_ASSESSMENT row — model outputs, derived flags ----
    fraud_assessments.append({
        'assessment_id':              f"FA-{str(uuid.uuid4())[:8].upper()}",
        'claim_id':                   claim_id,          # FK → Claim
        'fraud_score':                fraud_score,
        'fraud_pattern':              fraud_pattern,
        'rule_flag':                  rule_flag,
        'decline_reason':             decline_reason,
        'claim_outcome':              claim_outcome,
        'fraud_reported':             'Y' if is_fraud else 'N',
        'is_complex_claim':           'Y' if is_complex else 'N',
        # Computed policy-period flags stored here (not on raw claim)
        'incident_in_policy_period':  'Y' if incident_in_policy else 'N',
        'policy_status_at_incident':  policy_status_at_incident,
        'incident_near_boundary':     'Y' if incident_near_boundary else 'N',
    })

    # ---- DOCUMENTS ----
    # 1. Claim form / submission text
    sub_text = random.choice(SUBMISSION_TEMPLATES).format(
        inc_type=inc_type, inc_date=inc.strftime('%Y-%m-%d'),
        location=location, city=city, state=state,
        year=yr, make=make, model=model,
        severity=severity.lower(), policy=policy, authorities=auth
    )
    documents.append({
        'document_id':      f"DOC-{str(uuid.uuid4())[:8].upper()}",
        'claim_id':         claim_id,
        'document_type':    'claim_form',
        'format':           'PDF',
        'page_count':       random.randint(2, 5),
        'submission_date':  inc.strftime('%Y-%m-%d'),
        'document_text':    sub_text,
        'authenticity_flag': 'Verified' if random.random() > 0.1 else 'Requires Review',
    })

    # 2. Adjuster notes — stored as document rows, NOT on claim row
    if is_complex:
        num_notes = random.choices([1, 2, 3], weights=[0.5, 0.35, 0.15])[0]
        for _ in range(num_notes):
            note_date = inc + timedelta(days=random.randint(1, 30))
            tmpl = random.choice(ADJUSTER_NOTE_TEMPLATES)
            note = tmpl.format(
                name=fake.name(),
                description=random.choice([
                    'moderate frontal damage','significant side impact',
                    'rear-end collision damage','vandalism to driver side',
                    'water damage from flooding','theft-related damage',
                ]),
                action=random.choice([
                    'vehicle inspection','independent medical exam','witness interview',
                    'scene investigation','document review','SIU referral',
                ]),
                date=(note_date + timedelta(days=random.randint(3,14))).strftime('%Y-%m-%d'),
                location=random.choice(['Body Shop A','Dealership B',
                                         "Claimant's Residence",'Scene Location']),
                observations=random.choice([
                    'damage consistent with reported incident',
                    'additional pre-existing damage noted',
                    'damage appears more extensive than reported',
                    'damage pattern inconsistent with collision type',
                ]),
                shop=random.choice(['Precision Auto Body','City Collision Center',
                                     'Elite Repairs','Fast Fix Auto']),
                provider=random.choice(['Dr. Chen','Dr. Williams',
                                         'Metro Medical Center','Ortho Specialists']),
                findings=random.choice([
                    'cervical strain with limited ROM','lumbar contusion resolving',
                    'no objective findings','symptoms inconsistent with mechanism',
                ]),
                days=random.randint(7, 90),
                ime_status=random.choice(['scheduled','completed','pending','waived']),
                witness_name=fake.name(),
                consistency=random.choice(['corroborates','partially contradicts',
                                            'fully supports','differs from']),
                needed=random.choice(['required','not required','in progress']),
                amount=random.randint(1500, 15000),
                parts_status=random.choice(['OEM','Aftermarket','Backordered']),
                rental_status=random.choice(['approved','denied','pending','arranged']),
                count=random.randint(1, 5),
                years=random.randint(2, 10),
                pattern_status=random.choice(['identified','not evident','under review']),
                coverage_status='falls' if incident_in_policy else 'DOES NOT fall',
                limit_status='is' if total < 50000 else 'may not be',
                deductible=random.choice(DEDUCTIBLES),
                authority=random.choice(['Police Dept','Highway Patrol','Sheriff Office']),
                report_num=random.randint(10000, 99999),
                status=random.choice(['obtained','pending','unavailable']),
                match_status=random.choice(['matches','partially matches','contradicts']),
                doc_type=random.choice(['medical records','repair estimate',
                                         'wage loss statement','police report']),
                authenticity=random.choice(['legitimate','requires verification','appears altered']),
                discrepancy_note=random.choice([
                    '','Dates do not align with incident timeline.',
                    'Provider not in network.','Amounts appear inflated.',
                ]),
                additional_docs=random.choice(['medical authorization','proof of ownership',
                                                'wage verification','prior medical records']),
                evidence=random.choice(['scene photos','witness statements',
                                         'police report','video footage']),
                liability_pct=random.choice([0, 25, 50, 75, 100]),
                negligence_note=random.choice(['not a factor','under investigation','contributing factor']),
            )
            documents.append({
                'document_id':      f"DOC-{str(uuid.uuid4())[:8].upper()}",
                'claim_id':         claim_id,
                'document_type':    'adjuster_note',
                'format':           'TXT',
                'page_count':       1,
                'submission_date':  note_date.strftime('%Y-%m-%d'),
                'document_text':    note,
                'authenticity_flag': 'N/A',
            })

    # 3. Police report
    if auth not in ['None','Other']:
        documents.append({
            'document_id':      f"DOC-{str(uuid.uuid4())[:8].upper()}",
            'claim_id':         claim_id,
            'document_type':    'police_report',
            'format':           'PDF',
            'page_count':       random.randint(1, 3),
            'submission_date':  (inc + timedelta(days=random.randint(1,7))).strftime('%Y-%m-%d'),
            'document_text':    f"Police report #{random.randint(10000,99999)} - {inc_type} at {location}",
            'authenticity_flag': 'Verified' if random.random() > 0.05 else 'Requires Review',
        })

    # 4. Repair estimate
    if veh_cl > 0:
        if is_fraud and random.random() < 0.6:
            parts_total = round(veh_cl * random.uniform(0.55, 0.75), 2)
            labor_total = round(veh_cl * random.uniform(0.20, 0.35), 2)
            grand_total = round(veh_cl * random.uniform(1.05, 1.30), 2)
        else:
            parts_total = round(veh_cl * random.uniform(0.35, 0.55), 2)
            labor_total = round(veh_cl * random.uniform(0.25, 0.40), 2)
            grand_total = round(parts_total + labor_total * random.uniform(1.08, 1.12), 2)

        shop_name = random.choice(['Precision Auto Body','City Collision Center',
                                    'Elite Repairs','Fast Fix Auto','Dealer Collision Center'])
        est_date   = (inc + timedelta(days=random.randint(1, 10))).strftime('%Y-%m-%d')
        completion = (date.fromisoformat(est_date) + timedelta(days=random.randint(3,21))).strftime('%Y-%m-%d')
        damage_items = ['Front bumper replacement','Rear bumper repair','Left fender replacement',
                        'Right door repair','Hood replacement','Windshield replacement',
                        'Headlight assembly','Radiator support','Frame straightening','Paint and refinishing']
        sel_dmg = random.sample(damage_items, 3)
        costs   = [round(parts_total * random.uniform(0.15, 0.50), 2) for _ in sel_dmg]
        hours_v = random.randint(8, 40)
        rate_v  = random.choice([85, 95, 110, 125, 140])
        ded_v   = random.choice(DEDUCTIBLES)
        est_text = random.choice(REPAIR_ESTIMATE_TEMPLATES).format(
            shop_name=shop_name, est_date=est_date,
            year=yr, make=make, model=model,
            vin=fake.vin(), odometer=random.randint(15000, 180000),
            damage_1=sel_dmg[0], cost_1=costs[0],
            damage_2=sel_dmg[1], cost_2=costs[1],
            damage_3=sel_dmg[2], cost_3=costs[2],
            hours=hours_v, rate=rate_v,
            labor_cost=round(hours_v * rate_v, 2),
            parts_total=parts_total, labor_total=labor_total,
            tax=round(grand_total * 0.08, 2),
            grand_total=grand_total,
            completion_date=completion,
            notes=random.choice(['All parts OEM unless noted','Aftermarket parts used where available',
                                  'Additional damage may be found during disassembly','Customer authorized repairs']),
            estimator_name=fake.name(),
            supplies=round(grand_total * 0.03, 2),
            deductible=ded_v,
            customer_pay=min(ded_v, grand_total),
            valid_until=(date.fromisoformat(est_date) + timedelta(days=30)).strftime('%Y-%m-%d'),
        )
        documents.append({
            'document_id':      f"DOC-{str(uuid.uuid4())[:8].upper()}",
            'claim_id':         claim_id,
            'document_type':    'repair_estimate',
            'format':           'PDF',
            'page_count':       random.randint(1, 4),
            'submission_date':  est_date,
            'document_text':    est_text,
            'authenticity_flag': 'Requires Review' if (is_fraud and random.random() < 0.5) else 'Verified',
        })

    # 5. Medical report
    if injury > 0 or bodily_injuries_val > 0:
        if is_fraud and random.random() < 0.5:
            recovery_weeks   = random.randint(8, 26)
            pain_level       = random.randint(7, 10)
            disability_status = 'Temporarily Totally Disabled'
        else:
            recovery_weeks   = random.randint(2, 12)
            pain_level       = random.randint(3, 7)
            disability_status = random.choice(['Not Disabled','Temporarily Partially Disabled'])

        exam_date = (inc + timedelta(days=random.randint(1, 5))).strftime('%Y-%m-%d')
        med_text  = random.choice(MEDICAL_REPORT_TEMPLATES).format(
            patient_name=fake.name(),
            dob=rand_date(date(1950,1,1), date(2005,1,1)).strftime('%Y-%m-%d'),
            exam_date=exam_date,
            referring_md=random.choice(['Dr. Smith','Dr. Johnson','Dr. Lee','Self-referred']),
            accident_date=inc.strftime('%Y-%m-%d'),
            symptoms=random.choice(['neck pain and headache','lower back pain radiating to left leg',
                                     'right shoulder pain with limited motion','whiplash with cervical strain']),
            onset=random.choice(['Immediate','Within 24 hours','Gradual over 48 hours']),
            pain_level=pain_level,
            cervical_findings=random.choice(['Spasm and tenderness at C4-C6','Limited ROM with guarding',
                                              'Unremarkable','Paraspinal tenderness bilateral']),
            lumbar_findings=random.choice(['Tenderness at L4-L5','Full ROM, no tenderness',
                                            'Spasm with radicular symptoms','Unremarkable']),
            rom_findings=random.choice(['Decreased cervical rotation by 40%','Full ROM all planes',
                                         'Guarded with pain at endpoints','Within functional limits']),
            neuro_findings=random.choice(['Intact','Diminished sensation C6 distribution',
                                           'Reflexes 2+ symmetric','Mild weakness right grip']),
            imaging_type=random.choice(['X-ray','MRI','CT Scan','X-ray and MRI']),
            imaging_date=exam_date,
            imaging_findings=random.choice(['No acute fracture','Disc bulge at C5-C6',
                                              'Mild degenerative changes','Within normal limits',
                                              'Soft tissue swelling noted']),
            diagnosis_1=random.choice(['Cervical strain','Lumbar sprain',
                                        'Whiplash associated disorder','Concussion','Thoracic strain']),
            diagnosis_2=random.choice(['Cervicalgia','Muscle spasm','Post-traumatic headaches',
                                        'Radiculopathy','Myofascial pain']),
            treatment_1=random.choice(['Physical therapy 2x/week for 6 weeks','Chiropractic care 3x/week',
                                        'Prescription anti-inflammatory medication','Muscle relaxants as needed']),
            treatment_2=random.choice(['Home exercise program','Heat/ice therapy',
                                        'Ergonomic assessment','Massage therapy','Activity modification']),
            follow_up_weeks=random.choice([2, 4, 6, 8]),
            recovery_weeks=recovery_weeks,
            disability_status=disability_status,
            rtw_date=(date.fromisoformat(exam_date) + timedelta(weeks=recovery_weeks)).strftime('%Y-%m-%d'),
            provider_name=random.choice(['Dr. Michael Chen, MD','Dr. Sarah Williams, DO',
                                          'Dr. James Rodriguez, MD','Dr. Emily Park, DC']),
            provider_credentials=random.choice(['Board Certified - Orthopedic Surgery',
                                                  'Board Certified - Physical Medicine & Rehabilitation',
                                                  'Licensed Chiropractor','Board Certified - Neurology']),
            facility_name=random.choice(['Metro General Hospital','County Medical Center',
                                          'Urgent Care Plus','Regional Trauma Center']),
            arrival_time=f"{random.randint(10,23)}:{random.choice(['00','15','30','45'])}",
            triage_level=random.choice(['2 - Emergent','3 - Urgent','4 - Semi-urgent']),
            chief_complaint=random.choice(['Neck pain after MVC','Back pain and headache',
                                            'Right shoulder pain','Chest pain from seatbelt']),
            collision_detail=random.choice(['rear-end collision','side impact driver side',
                                              'head-on collision','rollover']),
            bp=f"{random.randint(110,160)}/{random.randint(60,95)}",
            hr=random.randint(60, 100),
            rr=random.randint(12, 20),
            o2=random.randint(95, 100),
            primary_survey='Intact',
            secondary_survey=random.choice(['Cervical spine tenderness, otherwise unremarkable',
                                              'Diffuse soft tissue tenderness','No acute findings']),
            procedures=random.choice(['None','IV placed, labs drawn','Splint applied','Wound care']),
            xray_result=random.choice(['No acute fracture','Cervical spine clearance pending','Negative']),
            ct_result=random.choice(['Not performed','Negative for acute intracranial hemorrhage','Pending']),
            lab_result=random.choice(['Within normal limits','Not indicated','Pending']),
            disposition=random.choice(['Discharged home','Admitted for observation','Transferred to tertiary care']),
            discharge_instructions=random.choice(['Follow up with PCP in 5-7 days','Return precautions given',
                                                    'Orthopedic follow-up recommended','Neurology follow-up as needed']),
            follow_up=random.choice(['PCP in 1 week','Orthopedics in 2 weeks','As needed']),
            attending_name=random.choice(['Dr. Robert Kim','Dr. Amanda Patel',
                                           'Dr. Thomas Greene','Dr. Lisa Zhang']),
        )
        documents.append({
            'document_id':      f"DOC-{str(uuid.uuid4())[:8].upper()}",
            'claim_id':         claim_id,
            'document_type':    'medical_report',
            'format':           'PDF',
            'page_count':       random.randint(2, 8),
            'submission_date':  exam_date,
            'document_text':    med_text,
            'authenticity_flag': 'Requires Review' if (is_fraud and random.random() < 0.4) else 'Verified',
        })

df_claims           = pd.DataFrame(claims)
df_fraud_assessment = pd.DataFrame(fraud_assessments)
df_documents        = pd.DataFrame(documents)
print(f"  Claim: {len(df_claims):,} rows, {len(df_claims.columns)} cols")
print(f"  Fraud_Assessment: {len(df_fraud_assessment):,} rows, {len(df_fraud_assessment.columns)} cols")
print(f"  Claim_Documents: {len(df_documents):,} rows, {len(df_documents.columns)} cols")

# ============================================================================
# 8. Quote_History
# ============================================================================
print("Building Quote_History ...")
quotes = []
for i, pnum_val in enumerate(policy_pool):
    n_quotes = random.choices([1, 2, 3], weights=[0.4, 0.4, 0.2])[0]
    cust = policy_customer_map[pnum_val]
    for _ in range(n_quotes):
        q_date = rand_date(date(2018,1,1), date(2024,6,1))
        af   = round(random.uniform(0.78, 1.70), 3)
        df_  = round(random.uniform(0.85, 1.90), 3)
        cf   = round(random.uniform(0.80, 1.45), 3)
        vf   = round(random.uniform(0.88, 1.55), 3)
        lf   = round(random.uniform(0.92, 1.40), 3)
        base = round(random.uniform(280, 1800), 2)
        mp   = round(random.uniform(0, 0.14), 3)
        sd   = round(random.uniform(0, 0.12), 3)
        final = round(base * af * df_ * cf * vf * lf * (1 - mp - sd), 2)
        accepted = random.choices(['Yes','No','Pending'], weights=[0.55, 0.30, 0.15])[0]
        counter  = round(final * random.uniform(0.80, 0.97), 2) if accepted == 'No' else None
        quotes.append({
            'quote_id':                  str(uuid.uuid4())[:8].upper(),
            'policy_number':             pnum_val,
            'customer_id':               cust,
            'quote_date':                q_date.strftime('%Y-%m-%d'),
            'quote_status':              random.choice(QUOTE_STATUS),
            'quote_accepted':            accepted,
            'counter_offer_premium':     counter,
            'risk_score':                round(random.uniform(100, 900), 1),
            'risk_class':                random.choice(RISK_CLASSES),
            'age_factor':                af,
            'driving_history_factor':    df_,
            'credit_factor':             cf,
            'vehicle_factor':            vf,
            'location_factor':           lf,
            'base_premium':              base,
            'multi_policy_discount':     mp,
            'safe_driver_discount':      sd,
            'final_premium':             final,
            'declared_vehicle_use':      random.choice(VEH_USES),
            'declared_annual_mileage':   random.randint(3000, 35000),
            'declared_prior_claims':     random.choices([0,1,2,3], [0.50,0.29,0.14,0.07])[0],
            'prior_insurance':           random.choice(INSURERS),
            'application_occupation':    random.choice(OCCUPATIONS),
            'application_income_range':  random.choice(INCOME_RNGS),
            'competitor_price_estimate': round(final * random.uniform(0.88, 1.18), 2),
            'renewal_flag':              random.choice(['New','Renewal','Re-quote']),
        })

df_quotes = pd.DataFrame(quotes)
if len(df_quotes) > 50000:
    df_quotes = df_quotes.sample(50000, random_state=42).reset_index(drop=True)
print(f"  Quote_History: {len(df_quotes):,} rows, {len(df_quotes.columns)} cols")

# ============================================================================
# 9. Customer_History
# ============================================================================
print("Building Customer_History ...")
history = []
for i in range(N):
    n_events = random.choices([0,1,2,3,4], weights=[0.45,0.30,0.14,0.07,0.04])[0]
    cust = policy_customer_map[policy_pool[i]]
    for _ in range(n_events):
        ev_date = rand_date(date(2015,1,1), date(2024,6,1))
        ev_type = random.choice(['Claim Filed','Policy Renewed','Quote Requested',
                                  'Address Change','Vehicle Changed','Payment Late','Cancellation'])
        history.append({
            'policy_number':          policy_pool[i],
            'customer_id':            cust,
            'event_date':             ev_date.strftime('%Y-%m-%d'),
            'event_type':             ev_type,
            'event_severity':         random.choice(SEVERITIES) if 'Claim' in ev_type else '',
            'event_amount':           random.randint(500, 60000) if 'Claim' in ev_type else 0,
            'days_since_policy_start': (ev_date - date(2018,1,1)).days,
        })

df_history = pd.DataFrame(history) if history else pd.DataFrame(
    columns=['policy_number','customer_id','event_date','event_type',
             'event_severity','event_amount','days_since_policy_start'])
print(f"  Customer_History: {len(df_history):,} rows")

# ============================================================================
# 10. Analytical / Reporting Views  (assembled via merge — not physical tables)
#     These are the "fat joined" sheets for Excel — read-only, derived.
#     In a real DB these would be SQL VIEWs, not stored tables.
# ============================================================================
print("Building analytical views ...")

# -- vw_fraud_model_input: everything a fraud model needs, joined at query time --
# Computes age_at_incident and months_as_customer from source dates
df_fraud_view = (
    df_claims
    .merge(df_fraud_assessment, on='claim_id', how='left')
    .merge(df_customer_master[['customer_id','date_of_birth','customer_since_date',
                                'gender','education_level','occupation','hobbies',
                                'relationship_status','credit_score',
                                'capital_gains','capital_loss']], on='customer_id', how='left')
    .merge(df_vehicle[['vehicle_id','make','model','year','market_value',
                        'annual_mileage','telematics_score']], on='vehicle_id', how='left')
    .merge(df_policy[['policy_number','effective_date','expiration_date',
                       'policy_type','policy_status','total_annual_premium',
                       'umbrella_limit']], on='policy_number', how='left')
)
# Derive age and tenure from dates
df_fraud_view['incident_date_dt']       = pd.to_datetime(df_fraud_view['incident_date'])
df_fraud_view['date_of_birth_dt']       = pd.to_datetime(df_fraud_view['date_of_birth'])
df_fraud_view['customer_since_date_dt'] = pd.to_datetime(df_fraud_view['customer_since_date'])
df_fraud_view['age_at_incident']        = ((df_fraud_view['incident_date_dt'] -
                                             df_fraud_view['date_of_birth_dt']).dt.days // 365).astype(int)
df_fraud_view['months_as_customer']     = ((df_fraud_view['incident_date_dt'] -
                                             df_fraud_view['customer_since_date_dt']).dt.days // 30).astype(int)
df_fraud_view.drop(columns=['incident_date_dt','date_of_birth_dt','customer_since_date_dt'], inplace=True)

# -- Customer_Claims --
df_customer_claims = df_claims.merge(
    df_customer_master[['customer_id','customer_name','date_of_birth',
                         'customer_address','customer_phone','customer_email']],
    on='customer_id', how='left'
)

# -- Customer_Policies --
df_customer_policies = df_policy.merge(
    df_customer_master[['customer_id','customer_name','gender',
                         'education_level','occupation','credit_score',
                         'capital_gains','capital_loss']],
    on='customer_id', how='left'
).merge(
    df_vehicle[['vehicle_id','make','model','year','market_value']],
    on='vehicle_id', how='left'
)

# -- Customer_Quotes --
df_customer_quotes = df_quotes.merge(
    df_customer_master[['customer_id','customer_name']], on='customer_id', how='left'
)

# -- Customer_History_Detail --
df_customer_history = df_history.merge(
    df_customer_master[['customer_id','customer_name']], on='customer_id', how='left'
)

# -- Customer_Documents --
df_customer_documents = (
    df_documents
    .merge(df_claims[['claim_id','customer_id']], on='claim_id', how='left')
    .merge(df_customer_master[['customer_id','customer_name']], on='customer_id', how='left')
)

# ============================================================================
# 11. Customer_Aggregated_Dashboard
# ============================================================================
print("Building Customer_Aggregated_Dashboard ...")

claims_agg = df_claims.groupby('customer_id').agg(
    num_claims         =('claim_id',           'count'),
    total_claim_amount =('total_claim_amount',  'sum'),
    avg_claim_amount   =('total_claim_amount',  'mean'),
    total_injury_claim =('injury_claim',        'sum'),
    total_property_claim=('property_claim',     'sum'),
    total_vehicle_claim=('vehicle_claim',       'sum'),
).reset_index()

fraud_agg = df_fraud_assessment.merge(
    df_claims[['claim_id','customer_id']], on='claim_id', how='left'
).groupby('customer_id').agg(
    num_fraud_claims   =('fraud_reported',      lambda x: (x == 'Y').sum()),
    avg_fraud_score    =('fraud_score',         'mean'),
).reset_index()

policies_agg = df_policy.groupby('customer_id').agg(
    num_policies           =('policy_number',   'count'),
    total_annual_premium   =('total_annual_premium', 'sum'),
    active_policies        =('policy_status',   lambda x: (x == 'Active').sum()),
    avg_umbrella_limit     =('umbrella_limit',  'mean'),
).reset_index()

vehicle_agg = df_vehicle.groupby('customer_id').agg(
    avg_telematics_score   =('telematics_score','mean'),
).reset_index()

coverage_agg = (
    df_coverage
    .merge(df_policy[['policy_number','customer_id']], on='policy_number', how='left')
    .groupby('customer_id').agg(
        unique_coverage_codes=('coverage_code','nunique'),
    ).reset_index()
)

quotes_agg = df_quotes.groupby('customer_id').agg(
    num_quotes             =('quote_id',        'count'),
    quote_acceptance_rate  =('quote_accepted',  lambda x: (x == 'Yes').mean()),
    avg_final_premium      =('final_premium',   'mean'),
    avg_risk_score         =('risk_score',      'mean'),
).reset_index()

history_agg = df_history.groupby('customer_id').agg(
    num_events             =('event_type',      'count'),
    total_event_amount     =('event_amount',    'sum'),
).reset_index()

doc_agg = df_customer_documents.groupby('customer_id').agg(
    num_documents          =('document_id',     'count'),
    num_unique_doc_types   =('document_type',   'nunique'),
).reset_index()

df_dashboard = df_customer_master.copy()
for agg_df in [claims_agg, fraud_agg, policies_agg, vehicle_agg,
               coverage_agg, quotes_agg, history_agg, doc_agg]:
    df_dashboard = df_dashboard.merge(agg_df, on='customer_id', how='left')

numeric_cols = df_dashboard.select_dtypes(include=[np.number]).columns
df_dashboard[numeric_cols] = df_dashboard[numeric_cols].fillna(0)
df_dashboard['is_fraudulent_customer'] = (
    df_dashboard['num_fraud_claims'] > 0
).map({True: 'Yes', False: 'No'})

print(f"  Dashboard: {len(df_dashboard):,} rows, {len(df_dashboard.columns)} cols")

# ============================================================================
# 12. Write Excel
# ============================================================================
print("\nWriting Excel file ...")
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

wb = Workbook(write_only=False)
if 'Sheet' in wb.sheetnames:
    del wb['Sheet']

HEADER_FILL = PatternFill('solid', start_color='1F3864')
HEADER_FONT = Font(bold=True, color='FFFFFF', name='Arial', size=9)
DATA_FONT   = Font(name='Arial', size=9)

def write_sheet(wb, df, name, freeze='A2'):
    if df is None or df.empty:
        print(f"  Warning: '{name}' is empty, skipping.")
        return
    ws = wb.create_sheet(name)
    rows = list(dataframe_to_rows(df, index=False, header=True))
    for r_idx, row in enumerate(rows, 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(r_idx, c_idx, val)
            if r_idx == 1:
                cell.font      = HEADER_FONT
                cell.fill      = HEADER_FILL
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                cell.font      = DATA_FONT
                cell.alignment = Alignment(vertical='center',
                                           wrap_text=(c_idx == len(row)))
    for col in ws.columns:
        vals = [str(col[0].value or '')]
        for c in list(col)[1:201]:
            vals.append(str(c.value or ''))
        ws.column_dimensions[col[0].column_letter].width = min(max(len(v) for v in vals) + 2, 45)
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = freeze
    print(f"  Sheet '{name}' written — {len(df):,} rows x {len(df.columns)} cols")

# Core normalised tables
write_sheet(wb, df_customer_master,  'Customer_Master')
write_sheet(wb, df_vehicle,          'Vehicle')
write_sheet(wb, df_policy,           'Policy_Data')
write_sheet(wb, df_coverage,         'Coverage_Line')
write_sheet(wb, df_claims,           'Claim')
write_sheet(wb, df_fraud_assessment, 'Fraud_Assessment')
write_sheet(wb, df_documents,        'Claim_Documents')
write_sheet(wb, df_quotes,           'Quote_History')
write_sheet(wb, df_history,          'Customer_History')

# Analytical / reporting views (joined)
write_sheet(wb, df_fraud_view,           'vw_Fraud_Model_Input')
write_sheet(wb, df_customer_claims,      'vw_Customer_Claims')
write_sheet(wb, df_customer_policies,    'vw_Customer_Policies')
write_sheet(wb, df_customer_quotes,      'vw_Customer_Quotes')
write_sheet(wb, df_customer_history,     'vw_Customer_History_Detail')
write_sheet(wb, df_customer_documents,   'vw_Customer_Documents')
write_sheet(wb, df_dashboard,            'Customer_Aggregated_Dashboard')

out ='underwriting_50k_dataset.xlsx'
wb.save(out)
print(f"\nSaved → {out}")

# Summary
print("\n" + "="*60)
print("SCHEMA SUMMARY")
print("="*60)
print("Core tables (normalised — each column owned once):")
print(f"  Customer_Master  : {len(df_customer_master):>6,} rows  {len(df_customer_master.columns)} cols")
print(f"  Vehicle          : {len(df_vehicle):>6,} rows  {len(df_vehicle.columns)} cols")
print(f"  Policy_Data      : {len(df_policy):>6,} rows  {len(df_policy.columns)} cols")
print(f"  Coverage_Line    : {len(df_coverage):>6,} rows  {len(df_coverage.columns)} cols")
print(f"  Claim            : {len(df_claims):>6,} rows  {len(df_claims.columns)} cols")
print(f"  Fraud_Assessment : {len(df_fraud_assessment):>6,} rows  {len(df_fraud_assessment.columns)} cols")
print(f"  Claim_Documents  : {len(df_documents):>6,} rows  {len(df_documents.columns)} cols")
print(f"  Quote_History    : {len(df_quotes):>6,} rows  {len(df_quotes.columns)} cols")
print(f"  Customer_History : {len(df_history):>6,} rows  {len(df_history.columns)} cols")
print("\nAnalytical views (joined at query time — no data copied):")
print(f"  vw_Fraud_Model_Input        : {len(df_fraud_view):>6,} rows  {len(df_fraud_view.columns)} cols")
print(f"  vw_Customer_Claims          : {len(df_customer_claims):>6,} rows")
print(f"  vw_Customer_Policies        : {len(df_customer_policies):>6,} rows")
print(f"  Customer_Aggregated_Dashboard: {len(df_dashboard):>6,} rows")